import os
import logging
import re
from datetime import datetime
import sys
import tkinter as tk
from tkinter import messagebox, filedialog
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill,Font,Alignment
import win32com.client
import unicodedata

# Constants for consistent configuration
LOG_DIR = 'logs'
TIMESTAMP_FORMAT = '%Y%m%d_%H%M%S'
DATETIME_PATTERNS = [
    '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d %H:%M', '%Y/%m/%d %H:%M',
    '%Y-%m-%d', '%Y/%m/%d'
]
TIME_RANGE_SEPARATORS = ['〜', '～', '~']
SPEC_SHEETS = ["補助調書2","市内児童一覧","退所・受託児童一覧"]

YELLOW_FILL = PatternFill(patternType="solid", fgColor='FFFF00')
ORANGE_FILL  = PatternFill(patternType="solid", fgColor="FFD966") 

PINK_FILL = PatternFill(patternType="solid", fgColor='FFC0CB')



def setup_logging(debug_level='DEBUG'):
    """Configure logging with file and console output for debugging and tracking."""
    # Ensure logs directory exists
    os.makedirs(LOG_DIR, exist_ok=True)

    # Create timestamped log file
    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
    log_file = os.path.join(LOG_DIR, f'excel_comparison_{timestamp}.log')
    log_format = '%(asctime)s - %(levelname)s - %(message)s'

    # Map string debug level to logging level constant
    levels = {'DEBUG': logging.DEBUG, 'INFO': logging.INFO, 'WARNING': logging.WARNING}
    logging_level = levels.get(debug_level, logging.WARNING)

    # Configure logging with file and console handlers
    logging.basicConfig(
        level=logging_level,
        format=log_format,
        handlers=[
            logging.FileHandler(log_file, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    logging.info(f'Logging initialized at level: {debug_level}')


def create_root():
    """Create and hide a Tkinter root window for file dialogs."""
    root = tk.Tk()
    root.withdraw()  # Hide the window to prevent it from appearing
    return root


def select_directory(root, prompt):
    """Prompt user to select a directory using a dialog."""
    logging.debug(f'Showing directory selection dialog: {prompt}')
    messagebox.showinfo('情報', prompt)
    folder = filedialog.askdirectory(title=prompt)
    logging.info(f'Selected directory: {folder}' if folder else 'No directory selected')
    return folder


def show_message(title, message):
    """Display a message box with the specified title and message."""
    logging.debug(f'Message box - Title: {title}, Message: {message}')
    messagebox.showinfo(title, message)



def normalize_value(value):
    """Normalize cell values for consistent comparison, handling various data types."""
    # Handle empty or whitespace-only strings
    if value is None or (isinstance(value, str) and not value.strip()):
        return None

    # Return datetime objects as-is
    if isinstance(value, datetime):
        return value

    # Convert float to int if it's a whole number
    if isinstance(value, float) and value.is_integer():
        value = int(value)

    # Convert to string for further processing
    value = str(value).strip()

    # Normalize full-width to half-width (e.g., ３児 -> 3児, ＡＢＣ -> ABC)
    value = unicodedata.normalize('NFKC', value)

    # Remove unwanted characters
    value = value.replace('_x000D_', '').replace('\r', '').replace('\n', '')
    value = value.replace('"', '').replace(' ', '').replace('、', ',').replace('・', ',').replace('.', ',').replace('歳','')

    for word in ["理事長", "経営者", "園長"]:
        value = value.replace(word, '')

    # Return None for empty strings after cleaning
    if not value:
        return None

    # Treat specific time strings as empty
    if value in ['0', '0:00', '00:00:00', '12:00:00午前']:
        return None

    # Attempt to parse datetime strings
    for pattern in DATETIME_PATTERNS:
        try:
            return datetime.strptime(value, pattern)
        except ValueError:
            continue

    # Convert numeric strings to integers if possible
    try:
        num = float(value)
        if num.is_integer():
            return str(int(num))
    except ValueError:
        pass

    return value


def is_datetime_string(value):
    """Check if a string represents a valid datetime."""
    if not isinstance(value, str):
        return False
    for pattern in DATETIME_PATTERNS:
        try:
            datetime.strptime(value, pattern)
            return True
        except ValueError:
            continue
    return False


def extract_date(value):
    """Extract date part from datetime string or object."""
    if isinstance(value, datetime):
        return value.strftime('%Y/%m/%d')
    if not isinstance(value, str):
        return value
    for pattern in DATETIME_PATTERNS:
        try:
            return datetime.strptime(value, pattern).strftime('%Y/%m/%d')
        except ValueError:
            continue
    date_pattern = r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})'
    match = re.search(date_pattern, value)
    if match:
        year, month, day = match.groups()
        return f"{year}/{int(month):02d}/{int(day):02d}"
    return value


def normalize_time_range(time_str):
    """Normalize time range string to a consistent format."""
    if not isinstance(time_str, str):
        time_str = str(time_str).strip()
    # Standardize separators to '~'
    for sep in TIME_RANGE_SEPARATORS:
        time_str = time_str.replace(sep, '~')
    return time_str


def recalculate_excel(file_path):
    """Recalculate Excel formulas in the specified file using COM (Windows only)."""
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        workbook = excel.Workbooks.Open(os.path.abspath(file_path))
        workbook.RefreshAll()
        workbook.Save()
        workbook.Close()
        excel.Quit()
        logging.info(f"Recalculated formulas in {file_path}")
    except Exception as e:
        logging.error(f"Failed to recalculate {file_path}: {e}")
        raise
    
import logging

def add_headers(row, value, headers, ignore_list=None):
    if value and value not in headers.values() and (not ignore_list or value not in ignore_list):
        headers[row] = value
def get_row_headers(sheet1, sheet2, sheet_name):
    try:
        headers1, headers2 = {}, {}
        col_start = 1
        row_max = max(sheet1.max_row, sheet2.max_row)

        if sheet_name == "補助調書2":
            current_main_header_v1 = None
            current_main_header_v2 = None

            ignore_main_headers = ["常勤人数", "常勤換算数"]
            ignore_sub_headers = ["氏名"]

            for row in range(8, row_max + 1):
                # Read main & sub headers
                main_header_v1 = sheet1.cell(row=row, column=1).value
                main_header_v2 = sheet2.cell(row=row, column=1).value
                sub_header_v1 = sheet1.cell(row=row, column=5).value
                sub_header_v2 = sheet2.cell(row=row, column=5).value

                # Ignore unwanted sub headers
                if sub_header_v1 in ignore_sub_headers:
                    sub_header_v1 = None
                if sub_header_v2 in ignore_sub_headers:
                    sub_header_v2 = None

                # Update current main headers if found
                if main_header_v1:
                    if main_header_v1 in ignore_main_headers and current_main_header_v1:
                        combined1 = f"{current_main_header_v1} - {main_header_v1}"
                    else:
                        current_main_header_v1 = main_header_v1
                        combined1 = (
                            f"{current_main_header_v1} - {sub_header_v1}"
                            if sub_header_v1 else current_main_header_v1
                        )
                else:
                    combined1 = (
                        f"{current_main_header_v1} - {sub_header_v1}"
                        if sub_header_v1 else current_main_header_v1
                    )

                if main_header_v2:
                    if main_header_v2 in ignore_main_headers and current_main_header_v2:
                        combined2 = f"{current_main_header_v2} - {main_header_v2}"
                    else:
                        current_main_header_v2 = main_header_v2
                        combined2 = (
                            f"{current_main_header_v2} - {sub_header_v2}"
                            if sub_header_v2 else current_main_header_v2
                        )
                else:
                    combined2 = (
                        f"{current_main_header_v2} - {sub_header_v2}"
                        if sub_header_v2 else current_main_header_v2
                    )

                # Save headers
                add_headers(row, combined1, headers1)
                add_headers(row, combined2, headers2)

            col_start = 5

        elif sheet_name == "市内児童一覧":
            preload_headers = {5: "市内", 6: "市外", 7: "合計"}
            headers1.update(preload_headers)
            headers2.update(preload_headers)

            for row in range(10, row_max + 1):
                add_headers(row, sheet1.cell(row=row, column=2).value, headers1)
                add_headers(row, sheet2.cell(row=row, column=2).value, headers2)

            col_start = 2

        elif sheet_name == "退所・受託児童一覧":
            ignore_header = ["児童氏名"]

            for row in range(5, row_max + 1):
                add_headers(row, sheet1.cell(row=row, column=2).value, headers1, ignore_header)
                add_headers(row, sheet2.cell(row=row, column=2).value, headers2, ignore_header)

        # Find matching header pairs
        header_pairs = [
            (row1, row2)
            for row1, h1 in headers1.items()
            for row2, h2 in headers2.items()
            if h1 == h2
        ]

        # Paired values
        paired_values1 = set(headers1.values()) & set(headers2.values())
        paired_values2 = set(headers2.values()) & set(headers1.values())

        # Non-paired headers
        not_pair_headers1 = {row: h for row, h in headers1.items() if h not in paired_values1}
        not_pair_headers2 = {row: h for row, h in headers2.items() if h not in paired_values2}

        return header_pairs, headers1, headers2, not_pair_headers1, not_pair_headers2, col_start

    except Exception as e:
        logging.error(f"Error processing header pairs: {str(e)}")
        return [], {}, {}, {}, {}, 1


def compare_excel_files(file1_path, file2_path):
    """Compare two Excel files, highlight differences, and return results."""
    logging.info(f'Comparing files: {file1_path} vs {file2_path}')
    reports = []
    try:
        # Recalculate formulas in the second file
        recalculate_excel(file2_path)
        wb1 = openpyxl.load_workbook(file1_path, data_only=True)
        wb2 = openpyxl.load_workbook(file2_path, data_only=True)
        mismatch_count = 0

        # Map visible sheets by their titles
        sheets1 = {s.title: s.title for s in wb1.worksheets if s.sheet_state == 'visible'}
        sheets2 = {s.title: s.title for s in wb2.worksheets if s.sheet_state == 'visible'}

        # Find common sheets between both workbooks
        common_sheets = set(sheets1) & set(sheets2)

        if not common_sheets:
            logging.warning('No matching sheets found')
            show_message("警告", "同じ名前のシートが見つかりません。")
            return 'X', wb2, []

        for sheet_id in common_sheets:
            sheet_report = []
            sheet_mismatches = 0
            sheet1 = wb1[sheets1[sheet_id]]
            sheet2 = wb2[sheets2[sheet_id]]
            logging.info(f'Comparing sheets: {sheets1[sheet_id]} <-> {sheets2[sheet_id]}')

            # Determine maximum dimensions for comparison
            row_max = max(sheet1.max_row, sheet2.max_row)
            col_max = max(sheet1.max_column, sheet2.max_column)

            sheet_name = sheets1.get(sheet_id)


            if sheet_name in SPEC_SHEETS :
                header_pairs, headers1, headers2,not_pair_headers1,not_pair_headers2,col_start = get_row_headers(sheet1, sheet2, sheet_name)
                row_pairs = header_pairs
                
            else:
                row_pairs = [(row, row) for row in range(1, row_max + 1)]
                col_start = 1

            for row1, row2 in row_pairs:

                for col in range(col_start, 35):

                    if sheet_name == "市内児童一覧" and row2 > 9 and col == 14:
                        continue  # Skip column N (14) for rows > 9

                    try:

                        v1 = normalize_value(sheet1.cell(row1, col).value)
                        v2 = normalize_value(sheet2.cell(row2, col).value)
                        logging.debug(f'Cell ({row2}, {col}): {v1} vs {v2}')

                        if v1 is None and v2 is None:
                            continue

                        v1_str = str(v1).strip() if not isinstance(v1, datetime) else v1
                        v2_str = str(v2).strip() if not isinstance(v2, datetime) else v2

                        # Handle datetime comparisons
                        if isinstance(v1, datetime) or isinstance(v2, datetime) or is_datetime_string(
                                v1_str) or is_datetime_string(v2_str):
                            if extract_date(v1) != extract_date(v2):
                                sheet2.cell(row2, col).fill = PINK_FILL
                                mismatch_count += 1
                                sheet_mismatches += 1
                                sheet_report.append({
                                    'row1': row1, 'col1': col, 'val1': v1,
                                    'row2': row2, 'col2': col, 'val2': v2
                                })
                            continue

                        # Handle time range comparisons
                        if any(sep in str(v1) or sep in str(v2) for sep in TIME_RANGE_SEPARATORS):
                            t1 = normalize_time_range(str(v1))
                            t2 = normalize_time_range(str(v2))
                            if t1 != t2:
                                sheet2.cell(row2, col).fill = PINK_FILL
                                mismatch_count += 1
                                sheet_mismatches += 1
                                sheet_report.append({
                                    'row1': row1, 'col1': col, 'val1': v1,
                                    'row2': row2, 'col2': col, 'val2': v2
                                })
                            continue

                        # General comparison for other values
                        if v1_str != v2_str:
                            sheet2.cell(row2, col).fill = PINK_FILL
                            mismatch_count += 1
                            sheet_mismatches += 1
                            sheet_report.append({
                                'row1': row1, 'col1': col, 'val1': v1,
                                'row2': row2, 'col2': col, 'val2': v2
                            })

                    except Exception as e:
                        logging.error(f'Error at cell ({row2}, {col}): {e}')
                        mismatch_count += 1
                        continue

            if sheet_name in SPEC_SHEETS:
                if not_pair_headers1:
                    for row in not_pair_headers1:
                        v1 = normalize_value(sheet1.cell(row, 5).value)
                        if v1 is not None:
                            # Missing in sheet2
                            sheet2.cell(row, 5).fill = ORANGE_FILL
                            mismatch_count += 1
                            sheet_mismatches += 1
                            sheet_report.append({
                                'row1': row, 'col1': 5, 'val1': v1,
                                'row2': row, 'col2': 5, 'val2': "MISSING"
                            })

                if not_pair_headers2:
                    for row in not_pair_headers2:
                        v2 = normalize_value(sheet2.cell(row, 5).value)
                        if v2 is not None:
                            # Missing in sheet1
                            sheet2.cell(row, 5).fill = ORANGE_FILL
                            mismatch_count += 1
                            sheet_mismatches += 1
                            sheet_report.append({
                                'row1': row, 'col1': 5, 'val1': "MISSING",
                                'row2': row, 'col2': 5, 'val2': v2
                            })

            if sheet_report:
                reports.append({
                    "sheet_name": sheet_id,
                    "sheet_report": sheet_report,
                    "mismatch_found": sheet_mismatches
                })
                
        result = 'X' if mismatch_count > 0 else 'O'
        logging.info(f'Comparison result: {result} (mismatches: {mismatch_count})')
        return result, wb2, reports

    except Exception as e:
        logging.error(f'Comparison error: {e}', exc_info=True)
        raise


def generate_excel_report(all_reports,compare_folder):
    """Generate an Excel report summarizing comparison results."""
    wb = openpyxl.Workbook()
    summary_sheet = wb.active
    summary_sheet.title = "Summary"

    # Set up summary sheet headers
    headers = ["School ID", "Mismatch Count", "Status", "Files"]
    for col, header in enumerate(headers, 1):
        cell = summary_sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    summary_row = 2
    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)

    if not all_reports:
        summary_sheet.cell(row=summary_row, column=1).value = "No mismatches found"
        summary_sheet.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=4)
    else:
        for school_report in all_reports:
            for school_id, file_reports in school_report.items():
                # if not file_reports:
                #     continue
                school_sheet = wb.create_sheet(re.sub(r'[\\\/:*?"<>|]', '_', school_id)[:31])
                current_row = 1
                school_mismatch_total = 0

                for file_report in file_reports:
                    workbook_name = list(file_report.keys())[0]
                    school_sheet.cell(row=current_row, column=1).value = f"Workbook: {workbook_name}"
                    school_sheet.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 2

                    # Set up headers for detailed report
                    headers = ["Sheet Name", "Count", "V1 Row", "V1 Col", "V1 Value", "V2 Row", "V2 Col", "V2 Value"]
                    for col, header in enumerate(headers, 1):
                        school_sheet.cell(row=current_row, column=col).value = header
                        school_sheet.cell(row=current_row, column=col).font = Font(bold=True)
                        school_sheet.cell(row=current_row, column=col).alignment = Alignment(horizontal='center')
                    current_row += 1

                    for sheet_report in file_report[workbook_name]:
                        sheet_name = sheet_report["sheet_name"]
                        mismatches = sheet_report["sheet_report"]
                        mismatch_count = sheet_report["mismatch_found"]

                        if mismatch_count == 0:
                            school_sheet.cell(row=current_row, column=1).value = sheet_name
                            school_sheet.cell(row=current_row, column=2).value = mismatch_count
                            school_sheet.cell(row=current_row, column=3).value = "-"
                            school_sheet.cell(row=current_row, column=5).value = "No mismatches"
                            current_row += 1
                            continue

                        school_mismatch_total += mismatch_count
                        first_mismatch = True

                        for mismatch in mismatches:
                            school_sheet.cell(row=current_row, column=1).value = sheet_name if first_mismatch else ""
                            school_sheet.cell(row=current_row,
                                              column=2).value = mismatch_count if first_mismatch else ""
                            school_sheet.cell(row=current_row, column=3).value = mismatch["row1"]
                            school_sheet.cell(row=current_row, column=4).value = mismatch["col1"]
                            school_sheet.cell(row=current_row, column=5).value = mismatch["val1"]
                            school_sheet.cell(row=current_row, column=6).value = mismatch["row2"]
                            school_sheet.cell(row=current_row, column=7).value = mismatch["col2"]
                            school_sheet.cell(row=current_row, column=8).value = mismatch["val2"]
                            first_mismatch = False
                            current_row += 1

                    current_row += 3

                summary_sheet.cell(row=summary_row, column=1).value = school_id
                summary_sheet.cell(row=summary_row, column=2).value = school_mismatch_total
                summary_sheet.cell(row=summary_row, column=3).value = "OK" if school_mismatch_total == 0 else "NG"
                summary_sheet.cell(row=summary_row, column=4).value = len(file_reports)
                summary_row += 1

    # Adjust column widths for readability
    for sheet in wb:
        for col in sheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    report_path = os.path.join(compare_folder, f"{timestamp}_comparison_report.xlsx")
    wb.save(report_path)
    return report_path


def generate_report(all_reports,compare_folder):
    """Generate a markdown report summarizing comparison results."""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    report_lines = [
        "# Excel Comparison Report",
        f"**Generated on:** {timestamp}",
        ""
    ]
    final_report = [
        "# Final Result Report",
        "",
        "| School ID | Mismatch | Status | Files |",
        "|-----------|----------|--------|-------|"
    ]

    if not all_reports:
        report_lines.append("## No mismatches found")
        report_lines.append("No differences were detected during the comparison.")
    else:
        for school_report in all_reports:
            for school_id, file_reports in school_report.items():
                report_lines.append(f"## School ID: {school_id}")
                report_lines.append("")
                school_mismatch_found = 0

                for file_report in file_reports:
                    workbook_name = list(file_report.keys())[0]
                    report_lines.append(f"### Workbook: {workbook_name}")
                    report_lines.append("")
                    report_lines.append(
                        "| Sheet Name | Count | V1 Row | V1 Col | V1 Value | V2 Row | V2 Col | V2 Value |")
                    report_lines.append(
                        "|------------|-------|--------|--------|----------|--------|--------|----------|")

                    for sheet_report in file_report[workbook_name]:
                        sheet_name = sheet_report["sheet_name"]
                        mismatches = sheet_report["sheet_report"]
                        mismatch_count = sheet_report["mismatch_found"]

                        if mismatch_count == 0:
                            report_lines.append(
                                f"| {sheet_name} | {mismatch_count} | - | - | No mismatches | - | - | - |")
                            continue

                        school_mismatch_found += mismatch_count
                        first_mismatch = True

                        for mismatch in mismatches:
                            row = f"| {' ' if not first_mismatch else sheet_name} | {' ' if not first_mismatch else mismatch_count} | {mismatch['row1']} | {mismatch['col1']} | {mismatch['val1']} | {mismatch['row2']} | {mismatch['col2']} | {mismatch['val2']} |"
                            report_lines.append(row)
                            first_mismatch = False

                    report_lines.append("")

                final_report.append(
                    f"| {school_id} | {school_mismatch_found} | {'OK' if school_mismatch_found == 0 else 'NG'} | {len(file_reports)} |")

    report_lines.extend([""] + final_report)
    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)
    report_path = os.path.join(compare_folder, f"{timestamp}_comparison_report.md")
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))
    return report_path


def process_folder(compare_folder):
    """Process all subfolders in the recompare directory, comparing Excel files."""
    try:
        all_reports = []
        # Get all subfolders in the recompare directory
        subfolders = [f for f in os.listdir(compare_folder) if os.path.isdir(os.path.join(compare_folder, f))]
        logging.info(f'Found {len(subfolders)} subfolders')
        

        for subfolder in subfolders:
            subfolder_path = os.path.join(compare_folder, subfolder)
            v1_path = os.path.join(subfolder_path, 'V1')
            v2_path = os.path.join(subfolder_path, 'V2')
            result_path = os.path.join(subfolder_path, 'result')

            # Skip if V1 or V2 folders are missing
            if not (os.path.exists(v1_path) and os.path.exists(v2_path)):
                logging.warning(f'Skipping {subfolder}: V1 or V2 folder missing')
                continue

            # Create result folder if it doesn't exist
            os.makedirs(result_path, exist_ok=True)
            logging.info(f'Processing subfolder: {subfolder}')

            # Get Excel files from V1 and V2 folders
            v1_files = [f for f in os.listdir(v1_path) if f.endswith(('.xlsx', '.xls'))]
            v2_files = set(f.strip() for f in os.listdir(v2_path) if f.endswith(('.xlsx', '.xls')))

            school_reports = []
            for file_name in v1_files:
                base_name = os.path.splitext(file_name)[0]
                file2_name = f"{base_name} .xlsx" if f"{base_name} .xlsx" in v2_files else f"{base_name}.xls"
                if file2_name not in v2_files:
                    logging.warning(f'No matching file for {file_name} in V2')
                    continue

                file1 = os.path.join(v1_path, file_name)
                file2 = os.path.join(v2_path, file2_name)
                logging.info(f'Processing: {file_name} vs {file2_name}')

                try:
                    result, modified_wb, reports = compare_excel_files(file1, file2)
                    output_path = os.path.join(result_path, f"{result}_{base_name}.xlsx")
                    modified_wb.save(output_path)
                    logging.info(f'Saved result to: {output_path}')
                    if reports:
                        school_reports.append({file2_name: reports})
                except Exception as e:
                    logging.error(f'Error processing {file_name}: {e}')
                    show_message("Error", f"Error processing {file_name}: {str(e)}")
                    continue

            all_reports.append({subfolder: school_reports})

        if all_reports:
            generate_report(all_reports,compare_folder)
            generate_excel_report(all_reports,compare_folder)
            logging.info('Reports generated successfully')
        else:
            logging.info('No mismatches found')
            show_message("No mismatches found", "No differences detected.")

        return True

    except Exception as e:
        logging.error(f'Error in process_folder: {e}', exc_info=True)
        return False


def main():
    """Main function to run the Excel comparison program."""
    setup_logging('DEBUG')
    logging.info('Starting Excel comparison program')
    root = None

    try:
        root = create_root()
        compare_folder = select_directory(root, "比較するフォルダを選択してください")
        if not compare_folder:
            logging.warning('Folder selection cancelled')
            show_message("フォルダー選択", "フォルダーが選択されていません。終了します...")
            return

        logging.info(f'Selected recompare folder: {compare_folder}')
        show_message("比較を開始します", "比較プロセスを開始しています....")

        if process_folder(compare_folder):
            logging.info('Comparison completed successfully')
            show_message("比較が完了しました", "比較プロセスが完了しました.")
        else:
            show_message("エラー", "処理中にエラーが発生しました。ログを確認してください。")

    except Exception as e:
        logging.error(f'Unexpected error: {e}', exc_info=True)
        messagebox.showerror("エラーが発生しました", f"予期しないエラーが発生しました: {str(e)}")
    finally:
        logging.info('Program finished')
        if root:
            root.destroy()


if __name__ == "__main__":
    main()